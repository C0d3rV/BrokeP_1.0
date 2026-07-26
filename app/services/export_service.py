"""Assembles trade data and writes it to Excel/PDF.

Two copy types, since a client bill and a broker bill legitimately differ:
  - "client": full detail, includes service fee, Net P&L = gross - brokerage - service fee
  - "broker": no service fee column at all, Net P&L = gross - brokerage only
    (service fee is the house's own margin -- the broker has no business
    seeing it, and it's never subtracted from the broker-facing net figure)
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib import colors
from reportlab.lib.pagesizes import landscape, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import cm

from app.domain.calculations.pnl import net_pl

CLIENT_HEADERS = ["Client", "Agent", "Segment", "Symbol", "Qty",
                   "Entry Date", "Entry Price", "Exit Date", "Exit Price",
                   "Brokerage", "Service Fee", "Gross P&L", "Net P&L", "Status"]

BROKER_HEADERS = ["Client", "Agent", "Segment", "Symbol", "Qty",
                   "Entry Date", "Entry Price", "Exit Date", "Exit Price",
                   "Brokerage", "Gross P&L", "Net P&L", "Status"]


def _trade_row(trade, client_name: str, agent_name: str, copy_type: str):
    brokerage_total = round((trade.entry_brokerage or 0) + (trade.exit_brokerage or 0), 2)
    fee_total = round((trade.entry_service_fee or 0) + (trade.exit_service_fee or 0), 2)
    gross = round(trade.gross_pl, 2) if trade.gross_pl is not None else "-"

    

    if copy_type == "client":
        base = [client_name, trade.segment, trade.symbol, trade.quantity,
                    trade.entry_date, trade.entry_price, trade.exit_date or "-", trade.exit_price or "-",
                    brokerage_total]
        net = round(trade.net_pl, 2) if trade.net_pl is not None else "-"
        return base + [fee_total, gross, net, trade.status]
    else:  # broker -- service fee excluded entirely, net recomputed without it
        base = [client_name, agent_name, trade.segment, trade.symbol, trade.quantity,
                trade.entry_date, trade.entry_price, trade.exit_date or "-", trade.exit_price or "-",
                brokerage_total]
        if trade.gross_pl is not None:
            net = round(net_pl(trade.gross_pl, trade.entry_brokerage or 0, trade.exit_brokerage or 0), 2)
        else:
            net = "-"
        return base + [gross, net, trade.status]


def _headers(copy_type: str):
    return CLIENT_HEADERS if copy_type == "client" else BROKER_HEADERS


def export_trades_to_excel(trades, client_name_lookup, agent_name_lookup,
                            filepath: str, copy_type: str = "client") -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Trades"

    headers = _headers(copy_type)
    ws.append(headers)
    header_fill = PatternFill(start_color="3B5BDB", end_color="3B5BDB", fill_type="solid")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for trade in trades:
        row = _trade_row(trade, client_name_lookup(trade.client_id),
                          agent_name_lookup(trade.agent_id), copy_type)
        ws.append(row)

    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[col[0].column_letter].width = max_len + 3

    ws.freeze_panes = "A2"
    wb.save(filepath)


def export_trades_to_pdf(trades, client_name_lookup, agent_name_lookup,
                          filepath: str, copy_type: str = "client",
                          title: str = "BrokeP Report") -> None:
    doc = SimpleDocTemplate(filepath, pagesize=landscape(A4),
                             leftMargin=1.2 * cm, rightMargin=1.2 * cm)
    styles = getSampleStyleSheet()
    elements = [Paragraph(title, styles["Title"])]

    headers = _headers(copy_type)
    data = [headers] + [
        _trade_row(t, client_name_lookup(t.client_id), agent_name_lookup(t.agent_id), copy_type)
        for t in trades
    ]
    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#3B5BDB")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 7.5),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#D8DEE9")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F7FC")]),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))
    elements.append(table)
    doc.build(elements)